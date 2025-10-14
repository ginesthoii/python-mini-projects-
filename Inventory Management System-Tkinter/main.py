import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import os

class InventoryApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Inventory System")

        self.inventory = []
        self.item_id = 1  # Auto-increment ID
        self.search_var = tk.StringVar()
        self.editing_item_id = None

        # ====== Left Frame (Form Inputs) ======
        left_frame = tk.Frame(self.root, padx=10, pady=10)
        left_frame.pack(side=tk.LEFT, fill=tk.Y)

        tk.Label(left_frame, text="Item Code:").grid(row=0, column=0, sticky="w")
        self.code_entry = tk.Entry(left_frame)
        self.code_entry.grid(row=0, column=1, pady=5)

        tk.Label(left_frame, text="Item Name:").grid(row=1, column=0, sticky="w")
        self.name_entry = tk.Entry(left_frame)
        self.name_entry.grid(row=1, column=1, pady=5)

        tk.Label(left_frame, text="Quantity:").grid(row=2, column=0, sticky="w")
        self.qty_entry = tk.Entry(left_frame)
        self.qty_entry.grid(row=2, column=1, pady=5)

        tk.Label(left_frame, text="Price:").grid(row=3, column=0, sticky="w")
        self.price_entry = tk.Entry(left_frame)
        self.price_entry.grid(row=3, column=1, pady=5)

        self.add_btn = tk.Button(left_frame, text="Add Item", command=self.add_item)
        self.add_btn.grid(row=4, column=0, columnspan=2, pady=5)

        self.update_btn = tk.Button(left_frame, text="Update Item", command=self.update_item, state=tk.DISABLED)
        self.update_btn.grid(row=5, column=0, columnspan=2, pady=5)

        clear_btn = tk.Button(left_frame, text="Clear Fields", command=self.clear_form)
        clear_btn.grid(row=6, column=0, columnspan=2, pady=5)

        delete_btn = tk.Button(left_frame, text="Delete Item", command=self.delete_item)
        delete_btn.grid(row=7, column=0, columnspan=2, pady=5)

        # ====== Right Frame (Table & Search) ======
        right_frame = tk.Frame(self.root, padx=10, pady=10)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        search_frame = tk.Frame(right_frame)
        search_frame.pack(anchor="w", pady=5)
        tk.Label(search_frame, text="Search:").pack(side=tk.LEFT)
        search_entry = tk.Entry(search_frame, textvariable=self.search_var, width=30)
        search_entry.pack(side=tk.LEFT, padx=5)
        self.search_var.trace("w", lambda *args: self.dynamic_search())

        columns = ("id", "code", "name", "qty", "price")
        self.tree = ttk.Treeview(right_frame, columns=columns, show="headings")
        for col in columns:
            self.tree.heading(col, text=col.capitalize(), command=lambda c=col: self.sort_by_column(c, False))
            self.tree.column(col, width=100)
        self.tree.pack(fill=tk.BOTH, expand=True)

        self.tree.bind("<Double-1>", self.on_item_double_click)

        # Load existing data if available
        self.load_data()

    def add_item(self):
        code = self.code_entry.get().strip()
        name = self.name_entry.get().strip()
        qty = self.qty_entry.get().strip()
        price = self.price_entry.get().strip()

        if not code or not name or not qty.isdigit() or not price.replace('.', '', 1).isdigit():
            messagebox.showerror("Error", "Please enter valid data.")
            return

        item = {
            "id": self.item_id,
            "code": code,
            "name": name,
            "qty": int(qty),
            "price": float(price)
        }
        self.inventory.append(item)
        self.item_id += 1

        self.refresh_table()
        self.clear_form()
        self.save_data()

    def update_item(self):
        if self.editing_item_id is None:
            return

        code = self.code_entry.get().strip()
        name = self.name_entry.get().strip()
        qty = self.qty_entry.get().strip()
        price = self.price_entry.get().strip()

        if not code or not name or not qty.isdigit() or not price.replace('.', '', 1).isdigit():
            messagebox.showerror("Error", "Please enter valid data.")
            return

        for it in self.inventory:
            if it["id"] == self.editing_item_id:
                it["code"] = code
                it["name"] = name
                it["qty"] = int(qty)
                it["price"] = float(price)
                break

        self.refresh_table()
        self.clear_form()
        self.save_data()

    def delete_item(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select an item to delete.")
            return

        item = self.tree.item(selected[0], "values")
        item_id = int(item[0])
        item_name = item[2]  # Item Name column

        confirm = messagebox.askyesno(
            "Confirm Delete",
            f"Are you sure you want to delete item '{item_name}'?"
        )
        if confirm:
            self.inventory = [it for it in self.inventory if it["id"] != item_id]
            self.refresh_table()
            self.save_data()
            self.clear_form()

    def refresh_table(self, data=None):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for it in (data if data is not None else self.inventory):
            self.tree.insert("", tk.END, values=(it["id"], it["code"], it["name"], it["qty"], it["price"]))

    def clear_form(self):
        self.code_entry.delete(0, tk.END)
        self.name_entry.delete(0, tk.END)
        self.qty_entry.delete(0, tk.END)
        self.price_entry.delete(0, tk.END)

        self.editing_item_id = None
        self.add_btn.config(state=tk.NORMAL)
        self.update_btn.config(state=tk.DISABLED)

        # Clear selection highlight
        for item in self.tree.selection():
            self.tree.selection_remove(item)

    def on_item_double_click(self, event):
        selected = self.tree.selection()
        if not selected:
            return

        item = self.tree.item(selected[0], "values")
        self.editing_item_id = int(item[0])

        self.code_entry.delete(0, tk.END)
        self.code_entry.insert(0, item[1])

        self.name_entry.delete(0, tk.END)
        self.name_entry.insert(0, item[2])

        self.qty_entry.delete(0, tk.END)
        self.qty_entry.insert(0, item[3])

        self.price_entry.delete(0, tk.END)
        self.price_entry.insert(0, item[4])

        self.add_btn.config(state=tk.DISABLED)
        self.update_btn.config(state=tk.NORMAL)

    def dynamic_search(self):
        query = self.search_var.get().lower()
        if not query:
            self.refresh_table(self.inventory)
        else:
            filtered = [
                it for it in self.inventory
                if query in str(it["id"]).lower()
                or query in it["code"].lower()
                or query in it["name"].lower()
            ]
            self.refresh_table(filtered)

    def sort_by_column(self, col, descending):
        data = [(self.tree.set(child, col), child) for child in self.tree.get_children("")]
        try:
            data.sort(key=lambda t: float(t[0]), reverse=descending)
        except ValueError:
            data.sort(key=lambda t: t[0].lower(), reverse=descending)

        for index, (val, child) in enumerate(data):
            self.tree.move(child, "", index)

        self.tree.heading(col, command=lambda: self.sort_by_column(col, not descending))

    def save_data(self):
        """Auto save inventory to Excel file with headers"""
        try:
            file_path = "inventory.xlsx"
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                ws = wb.active
                ws.delete_rows(2, ws.max_row)  # clear old data but keep headers
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Inventory"
                headers = ["ID", "Item Code", "Item Name", "Quantity", "Price"]
                ws.append(headers)

            for it in self.inventory:
                ws.append([it['id'], it['code'], it['name'], it['qty'], it['price']])

            wb.save(file_path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save data: {e}")

    def load_data(self):
        """Load inventory from Excel file if exists"""
        try:
            file_path = "inventory.xlsx"
            if not os.path.exists(file_path):
                return

            wb = load_workbook(file_path)
            ws = wb.active
            self.inventory.clear()
            self.item_id = 1

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                item = {
                    "id": int(row[0]),
                    "code": row[1],
                    "name": row[2],
                    "qty": int(row[3]),
                    "price": float(row[4])
                }
                self.inventory.append(item)
                self.item_id = max(self.item_id, item["id"] + 1)

            self.refresh_table()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load data: {e}")


if __name__ == "__main__":
    root = tk.Tk()
    app = InventoryApp(root)
    root.mainloop()
